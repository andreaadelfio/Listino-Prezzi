from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path("/home/andrea-adelfio/Dropbox/Documenti/Finanza/Portafogli/Portafogli.xlsx")
DEFAULT_OUTPUT = Path("/home/andrea-adelfio/Dropbox/Progetti/Python/portafogli-web/supabase/seed.sql")


def sql_quote(value: str | None) -> str:
    if value is None:
        return "null"
    return "'" + value.replace("'", "''") + "'"


def parse_price(price_text: str | None) -> tuple[str, str]:
    import re

    text = str(price_text or "").strip()
    match = re.search(r"(-?\d+(?:[.,]\d+)?)\s*(?:EUR|€)?(?:\s*/\s*([a-zA-Z]+))?", text, re.IGNORECASE)
    if not match:
        return "null", "null"
    value = match.group(1).replace(",", ".")
    unit = match.group(2)
    return value, sql_quote(unit) if unit else "null"


def generate_seed_sql(workbook_path: Path = DEFAULT_WORKBOOK, output_path: Path = DEFAULT_OUTPUT) -> None:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Listino Prezzi raw"]

    retailers: set[str] = set()
    rows: list[dict[str, str | None]] = []

    for row in range(2, ws.max_row + 1):
        prodotto = str(ws[f"A{row}"].value or "").strip()
        rivenditore = str(ws[f"B{row}"].value or "").strip()
        categoria = str(ws[f"D{row}"].value or "").strip()
        prezzo = str(ws[f"E{row}"].value or "").strip()

        if not prodotto or len(prodotto) <= 1:
            continue
        if not rivenditore or not prezzo:
            continue

        if rivenditore:
            retailers.add(rivenditore)

        rows.append(
            {
                "prodotto": prodotto,
                "rivenditore": rivenditore or None,
                "categoria": categoria or None,
                "prezzo": prezzo or None,
            }
        )

    lines: list[str] = []
    lines.append("-- Seed generato automaticamente da Listino Prezzi raw")
    lines.append("truncate table public.listino_prezzi_raw restart identity cascade;")
    lines.append("truncate table public.retailers restart identity cascade;")
    lines.append("")

    if retailers:
        lines.append("insert into public.retailers (name) values")
        retailer_values = ",\n".join(f"  ({sql_quote(name)})" for name in sorted(retailers))
        lines.append(retailer_values + ";")
        lines.append("")

    for item in rows:
        price_value, price_unit = parse_price(item["prezzo"])
        retailer_sql = (
            f"(select id from public.retailers where name = {sql_quote(item['rivenditore'])})"
            if item["rivenditore"]
            else "null"
        )
        lines.append(
            "insert into public.listino_prezzi_raw "
            "(prodotto, retailer_id, categoria, prezzo, prezzo_valore, prezzo_unita) values "
            f"({sql_quote(item['prodotto'])}, {retailer_sql}, {sql_quote(item['categoria'])}, "
            f"{sql_quote(item['prezzo'])}, {price_value}, {price_unit});"
        )

    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    generate_seed_sql()
